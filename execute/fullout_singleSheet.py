import win32gui,win32con,win32api,win32ui
# import re
import pyautogui
import re, traceback
import time
import datetime
import sys
import tempfile
import argparse
import os

import xlsxwriter

from PIL import Image
import pytesseract
import cv2

url = 'http://192.168.10.20:8004/api'
# url = 'http://127.0.0.1:8000/api'
x = 0
y = 0 
w = 0
h = 0 
vHeighScreen = 0
vBookingCreatePage = False
vCuurentBookingMode = ''
x_capture =0
y_capture = 0
w_capture = 0
h_capture = 0

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'


class WindowMgr:
    """Encapsulates some calls to the winapi for window management"""
    settingFile =''


    def __init__ (self):
        """Constructor"""
        self.hwnd = None

    def find_window(self,title):
        try:
            self.hwnd = win32gui.FindWindow(None, title)
            assert self.hwnd
            return self.hwnd
        except:
            pyautogui.alert(text='Not found program name ' + title + '\n' 
                            'Please open program before excute script', title='Unable to open program', button='OK')
            # print ('Not found program')
            return None


    def set_onTop(self,hwnd):
        win32gui.SetForegroundWindow(hwnd)
        return win32gui.GetWindowRect(hwnd)



    def Maximize(self,hwnd):
        win32gui.ShowWindow(hwnd,win32con.SW_RESTORE)#, win32con.SW_MAXIMIZE

    def get_mouseXY(self):
        return win32gui.GetCursorPos()

    def set_mouseXY(self):
        import os.path
        import json
        x,y,w,h = win32gui.GetWindowRect(self.hwnd)
        print ('Current Window X : %s  Y: %s' %(x,y))
        fname = 'setting.json'
        if os.path.isfile(fname) :
            dict = eval(open(fname).read())
            x1 = dict['x']
            y1 = dict['y']
            print ('Setting X : %s  Y: %s' %(x1,y1))
        win32api.SetCursorPos((x+x1,y+y1))
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, x+x1, y+y1, 0, 0)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, x+x1, y+y1, 0, 0)
        print ('Current Mouse X %s' % self.get_mouseXY()[0])
        print ('Current Mouse Y %s' % self.get_mouseXY()[1])


    def saveFirstDataPos(self):
        x,y,w,h = win32gui.GetWindowRect(self.hwnd)
        print ('Window X : %s  Y: %s' %(x,y))
        x1,y1 = self.get_mouseXY()
        print ('Mouse X : %s  Y: %s' %(x1,y1))
        data={}
        data['x'] = x1-x
        data['y'] = y1-y
        # f = open("setting.json", "w")
        # self.settingFile
        f = open(settingFile, "w")
        f.write(str(data))

        f.close()

    def wait(self,seconds=1,message=None):
        """pause Windows for ? seconds and print
an optional message """
        win32api.Sleep(seconds*1000)
        if message is not None:
            win32api.keybd_event(message, 0,0,0)
            time.sleep(.05)
            win32api.keybd_event(message,0 ,win32con.KEYEVENTF_KEYUP ,0)

    def typer(self,stringIn=None):
        PyCWnd = win32ui.CreateWindowFromHandle(self.hwnd)
        for s in stringIn :
            if s == "\n":
                self.hwnd.SendMessage(win32con.WM_KEYDOWN, win32con.VK_RETURN, 0)
                self.hwnd.SendMessage(win32con.WM_KEYUP, win32con.VK_RETURN, 0)
            else:
                print ('Ord %s' % ord(s))
                PyCWnd.SendMessage(win32con.WM_CHAR, ord(s), 0)
        PyCWnd.UpdateWindow()

    def WindowExists(windowname):
        try:
            win32ui.FindWindow(None, windowname)

        except win32ui.error:
            return False
        else:
            return True



def main():
    import json
    import sys
    from colorama import init, AnsiToWin32,Fore, Back, Style
    import math
    # import datetime
    init(wrap=False)
    stream = AnsiToWin32(sys.stderr).stream

    try:
        ldir = tempfile.mkdtemp()
        parser = argparse.ArgumentParser()
        parser.add_argument('-d', '--directory', default=ldir)
        args = parser.parse_args()
        tmpDir = args.directory
        # print (tmpDir)

        fname = 'fullout_setting.json'

        regex = "Session A - [24 x 80]"
        state_left = win32api.GetKeyState(0x01)  # Left button down = 0 or 1. Button up = -127 or -128
        state_right = win32api.GetKeyState(0x02)  # Right button down = 0 or 1. Button up = -127 or -128

        import os.path
        secs_between_keys=0.01

        win = WindowMgr()
        win_ins = win.find_window(regex)

       
        print (Style.RESET_ALL + '===Starting Process==', file=stream)
        if win_ins == None :
            print (Fore.RED + ('-Not found session : %s' % regex ), file=stream)
            print (Fore.RED + '-Please open CTCS program before execute..', file=stream)
            sys.exit()

        win.Maximize(win_ins)

        (x,y,w,h) =win.set_onTop(win_ins)
      
        #  Initial -Setup
        pyautogui.press('f12')
        pyautogui.press('f12')
        pyautogui.press('f12')
        pyautogui.press('f12')
        #2) Now On "Product Environment LCB1" screen.
        #Need to input "1" --> Work with CTCS.
        pyautogui.typewrite('1', interval=secs_between_keys)
        pyautogui.press('enter')

        #3) Now On "Select one of following" screen.
        #Need to input "1" --> Order.
        pyautogui.typewrite('1', interval=secs_between_keys)
        pyautogui.press('enter')

        #3) Now On "CTS order" screen.
        #Need to input "2" -->  Full Out Order
        pyautogui.typewrite('2', interval=secs_between_keys)
        pyautogui.press('enter')

        # ====Start VGM ======
        from openpyxl import load_workbook
        import io
        import glob
        import re
        rex = re.compile("^[A-Z]{4}[0-9]{7}$")


        txt_list = glob.glob('d:\\fullout\\*.txt')
        if len(txt_list) > 0 :
        	convert_txt_to_excel(txt_list[0])
        	# print('Found Text file : %s' % txt_list[0])

        file_list = glob.glob('d:\\fullout\\*.xlsx')
        print (Fore.GREEN + ('==============================================' ), file=stream)

        if len(file_list) == 0 :
            print (Fore.RED +'Not found any Excel file', file=stream)
            sys.exit()
        else:
            print ('Found %s' % file_list[0] )




        # sys.exit()


        for fullout_file in file_list:
            print (Fore.GREEN + ('File name : %s ' % fullout_file ), file=stream)                      
            with open(fullout_file, "rb") as f:
                in_mem_file = io.BytesIO(f.read())

            wb2 = load_workbook(in_mem_file, read_only=True,data_only=True)
            maxRow = 5000
            current_booking=''
            previous_booking =''
            first_booking = True
            for ws in wb2:
                print (Fore.GREEN + ('Sheet name : %s ' % ws.title ), file=stream)
                # Each sheet
                for index,row in enumerate(ws.iter_rows()) :
                    if index < 5 :
                        continue
                    if index == maxRow :
                        break

                    # booking = row[0].value.__str__().strip()
                    booking = row[12].value.__str__().strip()
                    current_booking = booking
                    container = row[1].value.__str__().strip()
                    line = row[11].value.__str__().strip()
                    agent = row[14].value.__str__().strip()
                    remark = row[15].value.__str__().strip()

                 

                    if not rex.match(container):
                        continue

                    if (previous_booking != current_booking) and not first_booking:
                        # go back to Booking page
                        pyautogui.press('enter',3)
                    

                    # # F6 to create new Document -- only first Container
                    if previous_booking != current_booking:
                        first_booking = False
                        print('New Booking %s' % current_booking)
                        create_order(booking,line,agent)
                        # sys.exit()

                    enter_container(container,remark)
                    print ('%s - %s - %s - %s - %s - %s' % (index,booking,container,line,agent,remark))
                    previous_booking = current_booking

                # pyautogui.press('enter',3)

        print ('======Finished=======')
        # ====End VGM=======
        sys.exit()

    except Exception as e:
    # except:
        if hasattr(e, 'message'):
            print(e.message)
        else:
            print(e)
        print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno), type(e).__name__, e)
        # f = open(tmpDir + "log.txt", "w")
        # f.write(traceback.format_exc())

def convert_txt_to_excel(text_file_name):
	head, tail = os.path.split(text_file_name)
	# print ('Convert file %s --%s' % (head,tail))
	with open(text_file_name) as f:
		content = f.readlines()
	content = [x.strip() for x in content]

	xls_name = '%s\\%s.xlsx' % (head,tail.split('.')[0])
	workbook = xlsxwriter.Workbook(xls_name)
	worksheet = workbook.add_worksheet('summary')
	worksheet.write(0, 1,'container')
	worksheet.write(0, 11,'line')
	worksheet.write(0, 12,'booking')
	worksheet.write(0, 14,'agent')
	worksheet.write(0, 15,'remark')

	idx = 0
	for c in content:
		if 'LINE =' in c :
			line = c.split('=')[1].strip()
		if 'AGENT =' in c :
			agent = c.split('=')[1].strip()
		if 'H02' in c :
			booking = c.split(' ')[0].replace('H02','')
		if 'H09' in c :
			container = c.split(' ')[0].replace('H09','')
			worksheet.write(idx+1, 1,container)
			worksheet.write(idx+1, 11,line)
			worksheet.write(idx+1, 12,booking)
			worksheet.write(idx+1, 14,agent)
			worksheet.write(idx+1, 15,'')
			idx=idx+1
	print ('Making Excel file completed')

	return xls_name

	


def create_order(booking,line,agent):
    secs_between_keys=0.01
    pyautogui.press('f6')
    pyautogui.typewrite(booking, interval=secs_between_keys)
    pyautogui.press('tab')
    pyautogui.typewrite(line, interval=secs_between_keys)
    pyautogui.press('tab')
    pyautogui.typewrite(agent, interval=secs_between_keys)
    pyautogui.press('enter')


def enter_container(container,remark):
    secs_between_keys=0.01
    pyautogui.typewrite(container, interval=secs_between_keys)
    pyautogui.press('tab',13)
    pyautogui.typewrite(remark, interval=secs_between_keys)
    pyautogui.press('tab',2)
    pyautogui.typewrite('Y', interval=secs_between_keys)
    pyautogui.press('enter')

def enter_vgm(liner,vgm):
    secs_between_keys=0.01
    pyautogui.press('down',5)
    pyautogui.press('delete',12)
    pyautogui.typewrite(vgm, interval=secs_between_keys) #Enter Gross Weight
    pyautogui.press('tab',2)
    pyautogui.press('delete',12) #Delete Weight Net
    pyautogui.press('tab',6)
    pyautogui.typewrite('Y', interval=secs_between_keys)
    pyautogui.press('enter')

def enter_extra_vgm():
    secs_between_keys=0.01
    pyautogui.typewrite('VGM', interval=secs_between_keys)
    pyautogui.press('tab',14)
    pyautogui.typewrite('12', interval=secs_between_keys)
    # Ready Y/N on Screen
    filename="images/vgm.png"
    im = pyautogui.screenshot(filename,region=(x_capture,y_capture,w_capture,h_capture))
    text = pytesseract.image_to_string(Image.open(filename), \
                config="--psm 6 --eom 3 -c tessedit_char_whitelist=-01234567890yXYZ:")
    print ('Reading value %s' % text)
    newVGM = False
    if ('N' in text) or ('I' in text):
        newVGM = True
    pyautogui.press('enter')
    return newVGM

def enter_verify_vgm(liner,vgm):
    secs_between_keys=0.01
    pyautogui.press('f6')
    pyautogui.typewrite(vgm, interval=secs_between_keys)
    pyautogui.press('tab',1)
    pyautogui.typewrite(liner, interval=secs_between_keys)
    pyautogui.press('enter')

def mofify_verify_vgm(liner,vgm):
    secs_between_keys=0.01
    pyautogui.typewrite('2', interval=secs_between_keys)
    pyautogui.press('enter')
    pyautogui.press('delete',9)
    pyautogui.typewrite(vgm, interval=secs_between_keys)
    pyautogui.press('enter')



main()

#2558