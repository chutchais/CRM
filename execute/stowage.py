import win32gui,win32con,win32api,win32ui
# import re
import pyautogui
import re, traceback
import time
import datetime
import sys
import tempfile
import argparse
import os

from PIL import Image
import pytesseract
import cv2

# Added  on Oct 5,2021
import redis
db = redis.StrictRedis('10.24.50.96', 6379,db=7, charset="utf-8", decode_responses=True) #Auto Berth Production

url = 'http://192.168.10.20:8004/api'
# url = 'http://127.0.0.1:8000/api'
x = 0
y = 0 
w = 0
h = 0 
vHeighScreen = 0
vBookingCreatePage = False
vCuurentBookingMode = ''
x_capture =0
y_capture = 0
w_capture = 0
h_capture = 0

x_status_capture = 0
y_status_capture = 0
w_status_capture = 0
h_status_capture = 0



class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'


class WindowMgr:
    """Encapsulates some calls to the winapi for window management"""
    settingFile =''


    def __init__ (self):
        """Constructor"""
        self.hwnd = None

    def find_window(self,title):
        try:
            self.hwnd = win32gui.FindWindow(None, title)
            assert self.hwnd
            return self.hwnd
        except:
            pyautogui.alert(text='Not found program name ' + title + '\n' 
                            'Please open program before excute script', title='Unable to open program', button='OK')
            # print ('Not found program')
            return None


    def set_onTop(self,hwnd):
        win32gui.SetForegroundWindow(hwnd)
        return win32gui.GetWindowRect(hwnd)



    def Maximize(self,hwnd):
        win32gui.ShowWindow(hwnd,win32con.SW_RESTORE)#, win32con.SW_MAXIMIZE

    def get_mouseXY(self):
        return win32gui.GetCursorPos()

    def set_mouseXY(self):
        import os.path
        import json
        x,y,w,h = win32gui.GetWindowRect(self.hwnd)
        print ('Current Window X : %s  Y: %s' %(x,y))
        fname = 'setting.json'
        if os.path.isfile(fname) :
            dict = eval(open(fname).read())
            x1 = dict['x']
            y1 = dict['y']
            print ('Setting X : %s  Y: %s' %(x1,y1))
        win32api.SetCursorPos((x+x1,y+y1))
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, x+x1, y+y1, 0, 0)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, x+x1, y+y1, 0, 0)
        print ('Current Mouse X %s' % self.get_mouseXY()[0])
        print ('Current Mouse Y %s' % self.get_mouseXY()[1])


    def saveFirstDataPos(self):
        x,y,w,h = win32gui.GetWindowRect(self.hwnd)
        print ('Window X : %s  Y: %s' %(x,y))
        x1,y1 = self.get_mouseXY()
        print ('Mouse X : %s  Y: %s' %(x1,y1))
        data={}
        data['x'] = x1-x
        data['y'] = y1-y
        # f = open("setting.json", "w")
        # self.settingFile
        f = open(settingFile, "w")
        f.write(str(data))

        f.close()

    def wait(self,seconds=1,message=None):
        """pause Windows for ? seconds and print
an optional message """
        win32api.Sleep(seconds*1000)
        if message is not None:
            win32api.keybd_event(message, 0,0,0)
            time.sleep(.05)
            win32api.keybd_event(message,0 ,win32con.KEYEVENTF_KEYUP ,0)

    def typer(self,stringIn=None):
        PyCWnd = win32ui.CreateWindowFromHandle(self.hwnd)
        for s in stringIn :
            if s == "\n":
                self.hwnd.SendMessage(win32con.WM_KEYDOWN, win32con.VK_RETURN, 0)
                self.hwnd.SendMessage(win32con.WM_KEYUP, win32con.VK_RETURN, 0)
            else:
                print ('Ord %s' % ord(s))
                PyCWnd.SendMessage(win32con.WM_CHAR, ord(s), 0)
        PyCWnd.UpdateWindow()

    def WindowExists(windowname):
        try:
            win32ui.FindWindow(None, windowname)

        except win32ui.error:
            return False
        else:
            return True


def container_executed(booking,container,skip=False):
    # Added on Oct 8,2021 --- To support SKIP checking excecuted.
    if skip :
        return False
    key = f'{booking}:{container}'
    return True if db.get(key) else False

def save_container(booking,container):
    key = f'{booking}:{container}'
    db.set(key,0)
    ttl = 60*60*24*7 #7 days
    db.expire(key, ttl)

def main():
    import json
    import sys
    from colorama import init, AnsiToWin32,Fore, Back, Style
    import math
    # import datetime
    init(wrap=False)
    stream = AnsiToWin32(sys.stderr).stream



    try:

        # Added on Oct 8,2021 -- To support SKIP verify execute mode.
        n = len(sys.argv)
        skip_mode = False
        if n > 1 :
            skip_mode = True if sys.argv[1]=='skip' else False 
        if skip_mode :
            print (Fore.RED + 'Running on SKIP verify executed mode........', file=stream)
            # print (f'Running on SKIP mode')

        # ldir = tempfile.mkdtemp()
        # parser = argparse.ArgumentParser()
        # parser.add_argument('-d', '--directory', default=ldir)
        # args = parser.parse_args()
        # tmpDir = args.directory
        # print (tmpDir)

        fname = 'vgm_setting.json'

        regex = "Session A - [24 x 80]"
        state_left = win32api.GetKeyState(0x01)  # Left button down = 0 or 1. Button up = -127 or -128
        state_right = win32api.GetKeyState(0x02)  # Right button down = 0 or 1. Button up = -127 or -128

        import os.path
        secs_between_keys=0.01

        win = WindowMgr()
        win_ins = win.find_window(regex)

        filename="images/vgm.png"

        # Finding Location of OCR command
        fOCR ='C:/Program Files (x86)/Tesseract-OCR/tesseract.exe'
        if os.path.exists(fOCR):
            print ('Using OCR command on %s' % fOCR )
            pytesseract.pytesseract.tesseract_cmd = fOCR

        fOCR ='C:/Program Files/Tesseract-OCR/tesseract.exe'
        if os.path.exists(fOCR):
            print ('Using OCR command on %s' % fOCR )
            pytesseract.pytesseract.tesseract_cmd = fOCR
        #===============================================
        
        print (Style.RESET_ALL + '===Starting Process==', file=stream)
        if win_ins == None :
            print (Fore.RED + ('-Not found session : %s' % regex ), file=stream)
            print (Fore.RED + '-Please open CTCS program before execute..', file=stream)
            sys.exit()

        win.Maximize(win_ins)
        global vHeighScreen
        global x
        global y
        global w
        global h
        global x_capture
        global y_capture
        global w_capture
        global h_capture

        global x_status_capture
        global y_status_capture
        global w_status_capture
        global h_status_capture

        # (x,y,w,h) =win.set_onTop(win_ins)
        (x,y,w,h) = win.set_onTop(win_ins)
        vHeighScreen = h*0.11

        # Container Status Setting -- using setting.json file
        fname_status= 'setting.json'
        if os.path.isfile(fname_status) :
            xx,yy,ww,hh = win32gui.GetWindowRect(win_ins)
            print (Style.RESET_ALL + 'Found Status Setting file', file=stream)
            dict = eval(open(fname_status).read())
            x2 = xx + dict['x']
            y2 = yy + dict['y']
            w2 = dict['w'] #- dict['x']
            h2 = dict['h']
        else:
            print (Style.RESET_ALL +'Not found Status Setting file' , file=stream)
            x2 = x
            y2 = h-vHeighScreen-5
            w2 = w-x
            h2 = vHeighScreen/2

        x_status_capture = x2
        y_status_capture = y2
        w_status_capture = w2
        h_status_capture = h2
        # print (xx,yy,ww,hh)
        # print (x_status_capture,y_status_capture,w_status_capture,h_status_capture)
        # filename_check='images/vgm_status.png'
        # im = pyautogui.screenshot(filename_check,region=(x_status_capture,y_status_capture,w_status_capture,h_status_capture))
        # text = pytesseract.image_to_string(Image.open(filename_check), \
        #             config="--psm 6 --eom 3 -c tessedit_char_whitelist=-01234567890yXYZ:")
        # print ('Reading Status value : %s' % text)

        # sys.exit()
        # -------------------------------------------------------------

        if os.path.isfile(fname) :
            print (Style.RESET_ALL + 'Found Setting file', file=stream)
            dict = eval(open(fname).read())
            x2 = x+dict['x']
            y2 = y+dict['y']
            w2 = dict['w']
            h2 = dict['h']
        else:
            print (Style.RESET_ALL +'Not found Setting file' , file=stream)
            x2 = x
            y2 = h-vHeighScreen-5
            w2 = w-x
            h2 = vHeighScreen/2

        x_capture = x2
        y_capture = y2
        w_capture = w2
        h_capture = h2



        # im = pyautogui.screenshot(filename,region=(x_capture,y_capture,w_capture,h_capture))
        # text = pytesseract.image_to_string(Image.open(filename), \
        #             config="--psm 6 --eom 3 -c tessedit_char_whitelist=-01234567890yXYZ:")
        # print ('Reading value : %s' % text)

       
        #  Initial -Setup
        pyautogui.press('f12')
        pyautogui.press('f12')
        pyautogui.press('f12')
        pyautogui.press('f12')
        #2) Now On "Product Environment LCB1" screen.
        #Need to input "1" --> Work with CTCS.
        pyautogui.typewrite('1', interval=secs_between_keys)
        pyautogui.press('enter')

        #3) Now On "Select one of following" screen.
        #Need to input "1" --> Order.
        pyautogui.typewrite('1', interval=secs_between_keys)
        pyautogui.press('enter')

        #3) Now On "CTS order" screen.
        #Need to input "1" -->  Booking (EMPTY OUT / FULL IN).
        pyautogui.typewrite('1', interval=secs_between_keys)
        pyautogui.press('enter')

        # ====Start VGM ======
        from openpyxl import load_workbook
        import io
        import glob
        file_list = glob.glob('d:\\stowage\\*.*')
        print (Fore.GREEN + ('==============================================' ), file=stream)

        if len(file_list) == 0 :
            print (Fore.RED +'Not found any Excel or Text file', file=stream)
            sys.exit()
        else:
            print ('Found %s' % file_list[0] )

        for stowage_file in file_list:
            print (Fore.GREEN + ('File name : %s ' % stowage_file ), file=stream)                      
            with open(stowage_file, "rb") as f:
                in_mem_file = io.BytesIO(f.read())

            # For Excel file
            if 'XLSX'  in stowage_file.upper() :
                wb2 = load_workbook(in_mem_file, read_only=True,data_only=True)
                maxRow = 500
                for ws in wb2:
                    print (Fore.GREEN + ('Sheet name : %s ' % ws.title ), file=stream)
                    # Each sheet
                    for index,row in enumerate(ws.iter_rows()) :
                        if index == maxRow :
                            break
                        if index == 0 :
                            print ('No data for sheet %s' % ws.title)
                            continue
                        booking = row[1].value.__str__().strip()
                        container = row[2].value.__str__().strip()
                        stowage = row[15].value.__str__().strip()

                        # Added on Oct 8,2021 -- To support SPOD
                        # For text file 100% to enter SPOD
                        spod = ''
                        discharge_port = row[8].value.__str__().strip()
                        if 'TPP' in  discharge_port :
                            spod = 'TPP'
                        
                        if 'SIN' in  discharge_port :
                            spod = 'SIN'
                        
                        # Added on Nov 25,2021 -- Support new SPOD
                        if 'NWK' in  discharge_port :
                            spod = 'NWK'


                        enter_booking_container_stwage(booking,container,stowage,spod,skip_mode)
                        # sys.exit()
                        # continue
                # Delete file
                delete_file (stowage_file)

                    # End Process#
            elif 'TXT'  in stowage_file.upper() : # For Text file
                # Using readlines()
                file1 = open(stowage_file, 'r')
                Lines = file1.readlines()
                file1.close()
                count = 0
                # Strips the newline character
                for line in Lines:
                    count += 1
                    if line =='':
                        continue
                    line_array = line.strip().split(' ')
                    container   = line_array[0]
                    if not is_valid_container(container):
                        continue
                    booking     = line_array[6]
                    stowage     = line_array[12][-2:]

                    # Added on Oct 8,2021 -- To support SPOD
                    # For text file 100% to enter SPOD
                    spod = line_array[12].strip()


                    if stowage == '' :
                        continue

                    # Added on Oct 12,2021 -- To support USNWKTM , no need to put spod
                    # except spod
                    if spod in ['NW2'] :
                        # spod = '' #Comment on Nov 25,2021 -- TO enable SPOD for NWK
                        stowage = 'N2'

                    # Start to process stowage
                    enter_booking_container_stwage(booking,container,stowage,spod,skip_mode)
                    # sys.exit()
                # Delete file
                delete_file (stowage_file)

        print ('======Finished=======')
        # ====End VGM=======
        sys.exit()

    except Exception as e:
    # except:
        if hasattr(e, 'message'):
            print(e.message)
        else:
            print(e)
        print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno), type(e).__name__, e)


def enter_booking_container_stwage(booking,container,stowage,spod='',skip=False):
    print (f'Process data {booking}/{container}/{stowage}.....start')
    if is_valid_container(container):
        if container_executed(booking,container,skip) : #check on Redis to verify Is it excuted.?
            print (f'{booking}/{container}......already excuted ..SKIP')
        else :
            # Check container is processed.
            enter_booking_container_list (booking)
            enter_container(container)
            # Added on Oct 8,2021 -- To support SPOD
            enter_stowage(stowage,spod)
            print (f'Process data {booking}/{container}/{stowage}.....finished')
            # Save to redis
            save_container(booking,container)
            time.sleep(1)
        # Save to Database.
    else:
        print (f'Process data {booking}/{container}/{stowage}.....Invalid container number SKIP..')

def is_valid_container(container) :
    import re
    rex = re.compile("^[A-Z]{4}[0-9]{7}$")
    return rex.match(container)

def delete_file(filename):
    import os
    os.remove(filename)
    print(f'Delete file {filename}')

def enter_booking_container_list(booking):
    secs_between_keys=0.01
    pyautogui.press('delete',20)
    pyautogui.typewrite(booking, interval=secs_between_keys)
    pyautogui.press('enter')
    pyautogui.typewrite('2', interval=secs_between_keys)
    pyautogui.press('enter')
    pyautogui.press('enter')

def enter_container(container):
    secs_between_keys=0.01
    pyautogui.hotkey('shift','f2')
    # pyautogui.press('up',2)
    # pyautogui.press('left',15)
    # Modify on Oct 4,2021 --
    pyautogui.press('up',3)
    pyautogui.press('tab',1)

    pyautogui.typewrite(container, interval=secs_between_keys)
    pyautogui.press('enter')
    pyautogui.typewrite('2', interval=secs_between_keys)
    pyautogui.press('enter')

# Added on Oct 4.2021 -- TO support Stowage file
def enter_stowage(stowage,spod=''):
    # SPOD is full name like TP1,TP2,SG1,SG2,TPP,SIN
    # Need to find index (COSMOS)
    
    secs_between_keys=0.01
    pyautogui.press('down',7)
    pyautogui.press('delete',2)
    pyautogui.typewrite(stowage, interval=secs_between_keys) #Enter stowage
    # pyautogui.press('enter')
    # move to unit sequence
    pyautogui.press('down',2)
    pyautogui.press('tab',2)
    pyautogui.press('delete',5) #delete existing seq
    # New Sequence number format hMMSS
    from datetime import datetime
    now = datetime.now()
    seq= now.strftime('%H%M%S')[-5:]
    seq = f'9{seq[1:]}' if seq[:1]=='0' else seq #To replace 0 with 9 ,in first digit
    
    pyautogui.typewrite(seq, interval=secs_between_keys)
    # Added on Oct 9,2021 -- To support SPOD
    if spod=='' :
        pyautogui.press('enter',3)
    else:
        # To enter spod , on SPOD window
        # Now after enter 5 digit for Seq ,cursor will move to OUT/IN 
        pyautogui.press('down',7) #to POD
        pyautogui.press('tab',1)    #To SPOD
        pyautogui.press('delete',2) #Delete SPOD
        pyautogui.press('enter',1)  #Open SPOD window
        # Select SPOD.
        # Find index of SPOD
        spod_ix = get_spod_index(spod)
        if spod_ix > 0 :
            pyautogui.press('down',spod_ix)
        pyautogui.typewrite('1', interval=secs_between_keys)
        pyautogui.press('enter',1) 
        pyautogui.press('enter',1)
        # Added on Oct 11,2021 -- Add more enter
        pyautogui.press('enter',1) 

def get_spod_index(spod):
    # return Zero based
    if spod == 'SG1':
        return 0
    if spod == 'SG2':
        return 1
    if spod == 'SIN':
        return 2

    if spod == 'TPP':
        return 0
    if spod == 'TP1':
        return 1
    if spod == 'TP2':
        return 2 
    if spod == 'TP3':
        return 3
    if spod == 'TP4':
        return 4
    
    # Added on Nov 25,2021 -- To support new POD
    if spod == 'NWK':
        return 0
    if spod == 'NW2':
        return 1
    
    return 0 #default


def enter_vgm(liner,vgm):
    secs_between_keys=0.01
    pyautogui.press('down',5)
    pyautogui.press('delete',12)
    pyautogui.typewrite(vgm, interval=secs_between_keys) #Enter Gross Weight
    pyautogui.press('tab',2)
    pyautogui.press('delete',12) #Delete Weight Net
    pyautogui.press('tab',6)
    pyautogui.typewrite('Y', interval=secs_between_keys)
    pyautogui.press('enter')
    # Add by Chutchai --To Handle In case Container problem.
    # Page will not move to next page
    filename="images/vgm_container_check.png"
    im = pyautogui.screenshot(filename,region=(x_status_capture,y_status_capture,w_status_capture,h_status_capture))
    text = pytesseract.image_to_string(Image.open(filename), \
                config="--psm 6 --eom 3 -c tessedit_char_whitelist=-01234567890yXYZ:")
    print ('Reading value %s' % text)

    if text != '' :
        print ('FOund problem on COntainer : press Shift+F2')
        pyautogui.hotkey('shift','f5')
    #----------------------------------------------------

def enter_extra_vgm(container_type):
    secs_between_keys=0.01
    pyautogui.typewrite('VGM', interval=secs_between_keys)
    if container_type=='RE':
        pyautogui.press('down',11)
    else:
        pyautogui.press('down',9)

    pyautogui.press('tab',1)
    # pyautogui.press('tab',14)

    pyautogui.typewrite('12', interval=secs_between_keys)
    # Ready Y/N on Screen
    filename="images/vgm.png"
    im = pyautogui.screenshot(filename,region=(x_capture,y_capture,w_capture,h_capture))
    text = pytesseract.image_to_string(Image.open(filename), \
                config="--psm 6 --eom 3 -c tessedit_char_whitelist=-01234567890yXYZ:")
    print ('Reading value %s' % text)
    newVGM = False
    if ('N' in text) or ('I' in text):
        newVGM = True
    pyautogui.press('enter')
    return newVGM

def enter_verify_vgm(liner,vgm):
    secs_between_keys=0.01
    pyautogui.press('f6')
    pyautogui.typewrite(vgm, interval=secs_between_keys)
    pyautogui.press('tab',1)
    pyautogui.typewrite(liner, interval=secs_between_keys)
    pyautogui.press('enter')

def mofify_verify_vgm(liner,vgm):
    secs_between_keys=0.01
    pyautogui.typewrite('2', interval=secs_between_keys)
    pyautogui.press('enter')
    pyautogui.press('delete',9)
    pyautogui.typewrite(vgm, interval=secs_between_keys)
    pyautogui.press('enter')



main()

#2558